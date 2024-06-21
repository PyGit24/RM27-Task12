import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service as ChromeService
import sys
import os

# Ensure the parent directory is in the system path for module imports
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from orange_login_page import OrangeLoginPage


@pytest.fixture
def browser():
    chrome_service = ChromeService(r"D:\ProgramS\PyCharm\chromedriver.exe")
    driver = webdriver.Chrome(service=chrome_service)
    driver.maximize_window()
    yield driver
    driver.quit()


def read_test_data_from_excel(
        path1=r"D:\ProgramS\PyCharm\Guvi\selenium\RM27\User_Login.xlsx", sheet_name='Sheet1'):
    wb = load_workbook(path1)
    ws = wb[sheet_name]

    test_data = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if row[2] is not None:
            test_data.append((row[1], row[2]))
        else:
            break

    return test_data, wb, ws


def test_orange_login(browser, username, password):
    login_page = OrangeLoginPage(browser)

    try:
        # Navigate to the Amazon login page
        login_page.navigate_to_login_page()

        # Wait for the sign-in button to be visible
        WebDriverWait(browser, 10).until(EC.visibility_of_element_located(login_page.sign_in_button))

        # Perform login
        login_page.click_sign_in_button()

        # Wait for username input field to be visible
        WebDriverWait(browser, 10).until(EC.visibility_of_element_located(login_page.username_locator))

        login_page.enter_username(username)
        login_page.continue_username()  # If "Continue" button is present, click it

        # Wait for password input field to be visible
        WebDriverWait(browser, 10).until(EC.visibility_of_element_located(login_page.password_locator))

        login_page.enter_password(password)
        login_page.click_login_button()

        # Wait for login success (customize as per your application)
        WebDriverWait(browser, 10).until(EC.title_contains("OrangeHRM"))
        return True
    except Exception as e:
        print(f"Error during login process: {e}")
        return False


@pytest.mark.parametrize("username,password", read_test_data_from_excel()[0])
def test_orange_login_data_driven(browser, username, password):
    test_data, wb, ws = read_test_data_from_excel()

    # Finding the row index for current username and password
    for row in ws.iter_rows(min_row=2):
        if row[1].value == username and row[2].value == password:
            row_index = row[1].row
            break

    if test_orange_login(browser, username, password):
        ws.cell(row=row_index, column=7).value = "Login successful"
    else:
        ws.cell(row=row_index, column=7).value = "Login unsuccessful"

    wb.save(r"D:\ProgramS\PyCharm\Guvi\selenium\RM27\User_Login.xlsx")


if __name__ == "__main__":
    pytest.main(["-v", "test_orange_login.py"])
